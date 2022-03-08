# Импорт модулей
from os import system
import openpyxl as xl

# Инструкция для терминала - очистка экрана
system('clear')

# Создание переменных
program_name = 'python excel merge'

# Создание списков
user_cols = []

# Создание словарей

# Создание функций
def print_program_name(program_name):
    """Функция вывода названия программы"""
    print ('==========================')
    print (program_name.upper())
    print ('==========================\n')

def get_excel_file():
    """Присваивание файла переменной"""
    file_name = input ('Введите имя файла: ')
    my_file = xl.load_workbook(file_name + '.xlsx')
    active_sheet = my_file.active
    return active_sheet

def create_dictionary(active_sheet):
    """Здесь происходит запись таблицы в словарь"""

    # Цикл для присваивания имен ключам значениями из первого столбца,
    #  т.к. первый столбец - артикул, и служит в качестве идентификатора.
    for column in active_sheet.columns:
        for cell in column:
            if (str(cell.value) != 'None'):
                table[cell.value] = ''
        break

    row_in_table = []
    rows_in_table = []
    # Цикл, который построчно заполняет список, который добавляется
    #  в новый список, перед этим удалив первое значение списка,
    #  т.к. эти значения в ключах словаря table, добавленные в
    #  функции create_dictionary.
    for row in active_sheet.rows:
        for cell in row:
            row_in_table.append(cell.value)
        row_in_table.pop(0)
        rows_in_table.append(row_in_table)
        row_in_table = []

    # Цикл для присваивания значений ключам (артикулам)
    circle = 0
    for key in table.keys():
        table[key] = rows_in_table[circle]
        circle += 1
    return table

def print_table(table):
    """Вывод словаря (таблицы)"""
    print ('\n')
    for key, values in table.items():
        print (str(key) + ': ' + str(values))

# =============================================

print_program_name(program_name)

tables = []
for circle in range(1,3):
    table = {}
    # Присваивание файла переменной
    active_sheet = get_excel_file()
    # Функция создания словаря
    table = create_dictionary(active_sheet)
    tables.append(table)
    
cir = 1
new_table = {}
circle = 0
# Если значение первой таблицы отсутствует во второй,
#  то добавить в новую таблицу
for key_one, row_one in tables[0].items():
    if (key_one not in tables[1].keys()):
        new_table[key_one] = row_one

# Если значение второй таблицы отсутствует в первой,
#  то добавить в новую таблицу        
for key_two, row_two in tables[1].items():
    if (key_two not in tables[0].keys()):
        new_table[key_two] = row_two

# Если ключ (артикул) первой таблицы равен ключу (артикулу) второй,
#  то суммировать значения записей обеих таблиц в третью,
#  в дальнейшем перезаписать table.
for key_one, row_one in tables[1].items():
    for key_two, row_two in tables[0].items():
        if (key_one == key_two):
            new_row = []
            sub_circle = 0
            while sub_circle < len(row_one):
                value = row_one[sub_circle] + row_two[sub_circle]
                new_row.append(value)
                new_table[key_one] = new_row
                sub_circle += 1
                cir += 1
table = new_table

# Объединение ключей и значений
#  таблицы в список записей.
table_rows = []
for key, values in table.items():
    values.insert(0, key)
    table_rows.append(values)

# Создание новой талицы
new_file = xl.Workbook()
# Задание названия первого листа
new_file.create_sheet(title = 'Сортировка', index = 0)
active_sheet = new_file['Сортировка']

# Вывод записей в таблицу
for table_row in table_rows:
    active_sheet.append(table_row)

# Сохранение нового файла
new_file.save('Отсортированная таблица.xlsx')

print ('Работа программы завершена.')
